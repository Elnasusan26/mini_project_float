from flask import (
    Flask, render_template, request,
    redirect, url_for, flash, session, abort, send_file
)
import io
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from collections import defaultdict
from functools import wraps
from datetime import datetime
from io import BytesIO
import pandas as pd

from models import (
    db, User, Class, Room, Subject,
    TimetableEntry, CancelledClass, TeachingAssignment,Teacher
)

from input_processor import process_inputs, process_lab_rooms
from allocator import allocate_rooms
from utils.normalize import normalize_slot

app = Flask(__name__)
app.secret_key = "floated-secret"

app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///database.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

db.init_app(app)

TIME_SLOTS = list(map(normalize_slot, [
    "8.00-8.45",
    "9.10-9.55",
    "10.00-10.45",
    "10.50-11.35",
    "11.55-12.40",
    "12.45-1.30"
]))

DAYS = [
    "MONDAY", "TUESDAY", "WEDNESDAY",
    "THURSDAY", "FRIDAY", "SATURDAY"
]

def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return wrapper


def role_required(role):
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if session.get("role") != role:
                abort(403)
            return f(*args, **kwargs)
        return wrapper
    return decorator


@app.route("/", methods=["GET", "POST"])
def login():

    if request.method == "POST":

        email = request.form.get("email")
        password = request.form.get("password")

        user = User.query.filter_by(email=email).first()

        if user and user.check_password(password):

            session.clear()
            session["user_id"] = user.id
            session["role"] = user.role

            if user.role == "admin":
                return redirect(url_for("admin_dashboard"))

            if user.role == "teacher":
                return redirect(url_for("teacher_dashboard"))

            if user.role == "student":
                return redirect(url_for("student_dashboard"))

        flash("Invalid email or password", "error")

    return render_template("login.html")
def get_cancelled_lookup(include_class_name=False):
    today = datetime.today().date()

    cancelled = CancelledClass.query.filter(
        CancelledClass.date >= today
    ).all()

    cancelled_lookup = set()

    for c in cancelled:
        cancel_day = c.date.strftime("%A").upper()
        slot = normalize_slot(c.slot)

        if include_class_name:
            cls = Class.query.get(c.class_id)
            if cls:
                cancelled_lookup.add((cls.name, cancel_day, slot))
        else:
            cancelled_lookup.add((c.class_id, cancel_day, slot))

    return cancelled_lookup

@app.route("/admin")
@login_required
@role_required("admin")
def admin_dashboard():

    permanent_count = Class.query.filter_by(
        class_category="permanent"
    ).count()

    floating_count = Class.query.filter_by(
        class_category="floating"
    ).count()

   
    from sqlalchemy import func, case

    subquery = (
        db.session.query(
            TimetableEntry.class_id,
            TimetableEntry.day,
            TimetableEntry.slot,
            func.count(TimetableEntry.id).label("total"),
            func.sum(
                case(
                    (TimetableEntry.room_id.isnot(None), 1),
                    else_=0
                )
            ).label("allocated")
        )
        .join(Class, TimetableEntry.class_id == Class.id)
        .filter(Class.class_category == "floating")
        .group_by(
            TimetableEntry.class_id,
            TimetableEntry.day,
            TimetableEntry.slot
        )
        .subquery()
    )

    allocated_count = (
        db.session.query(func.count())
        .filter(subquery.c.total == subquery.c.allocated)
        .scalar()
    )

    recent_cancelled = (
        CancelledClass.query
        .order_by(CancelledClass.id.desc())
        .limit(5)
        .all()
    )

    return render_template(
        "home.html",
        permanent_count=permanent_count,
        floating_count=floating_count,
        allocated_count=allocated_count,
        recent_cancelled=recent_cancelled
    )


@app.route("/admin_upload", methods=["GET", "POST"])
@login_required
@role_required("admin")
def admin_upload():

    if request.method == "POST":

        files = {
            "class_strength": "class_strength.xlsx",
            "room_mapping": "room_mapping.xlsx",
            "class_type": "class_type.xlsx",
            "teacher_subject": "teacher_subject_mapping.xlsx",
            "parallel_classes": "parallel_classes.xlsx",
            "student_mapping": "student_mapping.xlsx",
            "timetables": "timetables.xlsx",
            "lab_rooms": "lab_rooms.xlsx"
        }

        for key, filename in files.items():

            if key not in request.files or request.files[key].filename == "":
                return f"❌ Missing file: {key}", 400

            request.files[key].save(
                os.path.join(app.config["UPLOAD_FOLDER"], filename)
            )

        process_inputs()
        process_lab_rooms()
        allocate_rooms()

        return redirect(url_for("view_floating_timetable"))

    return render_template("admin_upload.html")


@app.route("/admin/cancel_class", methods=["GET", "POST"])
@login_required
@role_required("admin")
def cancel_class():

    classes = Class.query.all()

    if request.method == "POST":

        class_id = int(request.form.get("class_id"))
        date_str = request.form.get("date")
        date = datetime.strptime(date_str, "%Y-%m-%d").date()

        slots = request.form.getlist("slots")
        reason = request.form.get("reason")

        cls = Class.query.get(class_id)

        for slot in slots:

            slot = normalize_slot(slot)

            existing = CancelledClass.query.filter_by(
                class_id=class_id,
                slot=slot,
                date=date
            ).first()

            if existing:
                continue

            cancelled = CancelledClass(
                class_id=class_id,
                date=date,
                slot=slot,
                reason=reason
            )

            db.session.add(cancelled)
        db.session.commit()

        allocate_rooms()

        flash("Class cancelled and rooms reallocated!", "success")

        return redirect(url_for("cancelled_classes"))

    return render_template(
        "admin_cancel_class.html",
        classes=classes
    )

@app.route("/admin/cancelled_classes")
@login_required
@role_required("admin")
def cancelled_classes():

    cancelled = CancelledClass.query.order_by(
        CancelledClass.date.desc()
    ).all()

    return render_template(
        "cancelled_classes.html",
        cancelled=cancelled
    )


@app.route("/admin/delete_cancelled/<int:id>", endpoint="delete_cancelled")
@login_required
@role_required("admin")
def delete_cancelled(id):

    cancelled = CancelledClass.query.get_or_404(id)

    class_id = cancelled.class_id
    slot = normalize_slot(cancelled.slot)
    cancel_day = cancelled.date.strftime("%A").upper()

    db.session.delete(cancelled)
    db.session.commit()

    cls = Class.query.get(class_id)

    if cls and cls.class_category == "permanent":

        room = Room.query.filter_by(owner_class_id=class_id).first()

        if room:
            entries = TimetableEntry.query.filter_by(
                class_id=class_id,
                day=cancel_day,
                slot=slot
            ).all()

            for e in entries:
                e.room_id = room.id

            db.session.commit()

    allocate_rooms()

    return redirect(url_for("cancelled_classes"))
@app.route("/admin/faculty")
@login_required
@role_required("admin")
def faculty_list():
    teachers = Teacher.query.order_by(Teacher.name).all()

    # derive departments from classes each teacher teaches
    teacher_departments = {}
    for t in teachers:
        classes = (
            db.session.query(Class.name)
            .join(TimetableEntry, TimetableEntry.class_id == Class.id)
            .filter(TimetableEntry.teacher_id == t.id)
            .distinct()
            .all()
        )
        depts = set()
        for (cname,) in classes:
            # e.g. "S8_CSE" -> "CSE", "S6_CSE_A" -> "CSE"
            parts = cname.replace("-", "_").split("_")
            if len(parts) >= 2:
                last = parts[-1].upper()
                # if last part is a section letter, take the one before it
                dept = parts[-2].upper() if len(last) == 1 else last
                if len(dept) > 1:
                    depts.add(dept)
        teacher_departments[t.id] = sorted(depts)

    return render_template(
        "admin_faculty_list.html",
        teachers=teachers,
        teacher_departments=teacher_departments
    )

@app.route("/admin/faculty/<int:teacher_id>")
@login_required
@role_required("admin")
def faculty_timetable(teacher_id):

    teacher = Teacher.query.get_or_404(teacher_id)

    entries = (
        TimetableEntry.query
        .join(Subject)
        .join(Class)
        .filter(TimetableEntry.teacher_id == teacher.id)
        .order_by(TimetableEntry.day, TimetableEntry.slot)
        .all()
    )

    today = datetime.today().date()

    cancelled = CancelledClass.query.filter(
        CancelledClass.date >= today
    ).all()

    cancelled_lookup = set()

    for c in cancelled:
        cancel_day = c.date.strftime("%A").upper()
        slot = normalize_slot(c.slot)

        cancelled_lookup.add((c.class_id, cancel_day, slot))

    template = "admin_faculty_timetable.html" if session.get("role") == "admin" else "teacher_timetable.html"
    return render_template(
        "teacher_timetable.html",   
        entries=entries,
        cancelled_lookup=cancelled_lookup,
        teacher_name=teacher.name   
    )


@app.route("/view/timetable")
@app.route("/view/floating_timetable")
@app.route("/floating_timetable_grid")
@login_required
def view_floating_timetable():

    role = session.get("role")
    
    entries = TimetableEntry.query.order_by(
        TimetableEntry.class_id,
        TimetableEntry.day,
        TimetableEntry.slot
    ).all()

    raw = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))

    for e in entries:

        cls = e.class_obj.name
        day = e.day
        slot = normalize_slot(e.slot)

        raw[cls][day][slot].append({
            "subject": e.subject.name if e.subject else "-",
            "room": e.room.name if e.room else "-",
            "lab_rooms": e.lab_rooms,
            "batch": e.batch
        })

    cancelled_lookup = get_cancelled_lookup(include_class_name=True)
    class_map = {c.name: c.id for c in Class.query.all()}

    template = "floating_timetable_grid_teacher.html" if role == "teacher" else "floating_timetable_grid.html"

    return render_template(
        template,
        timetable=raw,
        slots=TIME_SLOTS,
        days=DAYS,
        cancelled_lookup=cancelled_lookup,
        class_map=class_map
    )

@app.route("/teacher")
@login_required
@role_required("teacher")
def teacher_dashboard():

    user = User.query.get(session["user_id"])

    if not user.teacher_id:
        return "Teacher account not linked to faculty record"

    entries = (
        TimetableEntry.query
        .join(TeachingAssignment,
            (TeachingAssignment.subject_id == TimetableEntry.subject_id) &
            (TeachingAssignment.class_id == TimetableEntry.class_id)
        )
        .filter(TeachingAssignment.teacher_id == user.teacher_id)
        .order_by(TimetableEntry.day, TimetableEntry.slot)
        .all()
    )

    for e in entries:
        print(e.subject.name if e.subject else None,
              e.class_obj.name if e.class_obj else None)

    cancelled_lookup = get_cancelled_lookup()

    return render_template(
        "teacher_timetable.html",
        entries=entries,
        cancelled_lookup=cancelled_lookup
    )

@app.route("/student")
@login_required
@role_required("student")
def student_dashboard():

    user = User.query.get(session["user_id"])

    cls = Class.query.get(user.class_id)

    entries = TimetableEntry.query.filter_by(
        class_id=user.class_id
    ).all()

    cancelled_lookup = {
    (c_day, slot)
    for (cid, c_day, slot) in get_cancelled_lookup()
    if cid == user.class_id
}

    return render_template(
        "student_timetable.html",
        entries=entries,
        class_name=cls.name,
        cancelled_lookup=cancelled_lookup
    )

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/class_timetable/<int:class_id>")
@login_required
def class_timetable(class_id):

    entries = TimetableEntry.query.filter_by(class_id=class_id)\
        .order_by(TimetableEntry.day, TimetableEntry.slot)\
        .all()

    return render_template(
        "class_timetable.html",
        entries=entries,
        class_id=class_id
    )
@app.route("/export_class_timetable/<int:class_id>")
@login_required
def export_class_timetable(class_id):

    cls = Class.query.get_or_404(class_id)

    wb = Workbook()
    ws = wb.active
    ws.title = cls.name

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bold = Font(bold=True)

    ws.merge_cells("A1:G1")
    ws["A1"] = "AISAT/Form/QPM18/F3"
    ws["A1"].alignment = center
    ws["A1"].font = bold

    ws.merge_cells("A2:G2")
    ws["A2"] = "Regular Class Timetable"
    ws["A2"].alignment = center
    ws["A2"].font = bold

    ws.merge_cells("A3:G3")
    ws["A3"] = f"Class: {cls.name}"
    ws["A3"].alignment = center

    ws.append([])

    headers = [
        "Day",
        "8.00 - 8.45",
        "9.10 - 9.55",
        "10.00 - 10.45",
        "10.50 - 11.35",
        "11.55 - 12.40",
        "12.45 - 1.30"
    ]

    ws.append(headers)

    entries = TimetableEntry.query.filter_by(class_id=class_id).all()

    data = {}

    for e in entries:
        data.setdefault(e.day, {})
        data[e.day].setdefault(e.slot, [])

        subject = e.subject.name if e.subject else "-"
        teacher = f"[{e.teacher.name}]" if e.teacher else ""

        if e.lab_rooms:
            room = f"({e.lab_rooms})"
        elif e.room:
            room = f"({e.room.name})"
        else:
            room = ""

        text = f"{subject}\n{teacher}\n{room}"

        data[e.day][e.slot].append(text)

    for day in DAYS:

        row = [day]

        for slot in TIME_SLOTS:

            cell = data.get(day, {}).get(slot, [])

            if cell:
                row.append("\n".join(cell))
            else:
                row.append("")

        ws.append(row)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = center

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name=f"{cls.name}_timetable.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
if __name__ == "__main__":

    with app.app_context():
        db.create_all()

    app.run(debug=True, use_reloader=False)